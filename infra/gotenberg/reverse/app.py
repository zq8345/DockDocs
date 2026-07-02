"""
DockDocs reverse-convert service
- /convert   : PDF → DOCX / XLSX  (server-side via Netlify proxy, ≤5 MB)
- /upload/convert : Office → PDF  (browser direct-upload, 5–100 MB, HMAC token)
Both run on port 4000 inside the Docker network; Caddy exposes them externally.
"""
import hashlib
import hmac
import io
import os
import tempfile
import time
import logging
from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Request
from fastapi.responses import Response
import httpx
import pdf2docx
import pdfplumber
import openpyxl

logging.basicConfig(level=logging.INFO)
app = FastAPI(docs_url=None, redoc_url=None)

SECRET = os.environ.get("DOCKDOCS_REVERSE_SECRET", "")
MAX_BYTES = 10 * 1024 * 1024  # 10 MB — larger than the 5 MB Netlify proxy cap for safety

SUPPORTED = {"pdf-to-word", "pdf-to-excel"}

# --- upload/convert constants -----------------------------------------------
GOTENBERG_INTERNAL_URL = os.environ.get("GOTENBERG_INTERNAL_URL", "http://gotenberg:3000")
UPLOAD_MAX_BYTES = 100 * 1024 * 1024  # 100 MB — full client direct-upload ceiling
TOKEN_WINDOW_SECS = 300               # 5-minute HMAC token validity
UPLOAD_ROUTES = {"word-to-pdf", "ppt-to-pdf", "excel-to-pdf", "html-to-pdf", "pdf-to-pdfa"}

# Default filenames when the browser doesn't supply one
_ROUTE_FILENAME = {
    "word-to-pdf": "document.docx",
    "ppt-to-pdf": "presentation.pptx",
    "excel-to-pdf": "spreadsheet.xlsx",
    "html-to-pdf": "index.html",
    "pdf-to-pdfa": "document.pdf",
}

# Gotenberg endpoint per route (internal network, no auth header needed)
def _gotenberg_endpoint(route: str) -> str:
    if route == "html-to-pdf":
        return f"{GOTENBERG_INTERNAL_URL}/forms/chromium/convert/html"
    if route == "pdf-to-pdfa":
        return f"{GOTENBERG_INTERNAL_URL}/forms/pdfengines/convert"
    return f"{GOTENBERG_INTERNAL_URL}/forms/libreoffice/convert"


def _verify_upload_token(token_header: str, route: str) -> bool:
    """Validate HMAC upload token: ts:nonce:hex(HMAC_SHA256(secret, 'ts:nonce:route'))"""
    if not SECRET or not token_header:
        return False
    try:
        ts_str, nonce, mac = token_header.split(":", 2)
        ts = int(ts_str)
        if abs(int(time.time()) - ts) > TOKEN_WINDOW_SECS:
            return False
        msg = f"{ts_str}:{nonce}:{route}".encode()
        expected = hmac.new(SECRET.encode(), msg, hashlib.sha256).hexdigest()
        return hmac.compare_digest(expected, mac)
    except Exception:
        return False


# ---------------------------------------------------------------------------
# /convert  — PDF → office (proxied through Netlify function, ≤5 MB)
# ---------------------------------------------------------------------------
@app.post("/convert")
async def convert(
    request: Request,
    route: str = Form(...),
    file: UploadFile = File(...),
):
    # Gate: shared secret (same pattern as Caddy → Gotenberg)
    if SECRET and request.headers.get("x-dockdocs-key") != SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")

    if route not in SUPPORTED:
        raise HTTPException(status_code=422, detail=f"Unsupported route: {route}. Supported: {sorted(SUPPORTED)}")

    data = await file.read()
    if len(data) > MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (max {MAX_BYTES // 1024 // 1024} MB).")
    if not data:
        raise HTTPException(status_code=422, detail="Empty file.")

    with tempfile.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, "source.pdf")
        with open(src, "wb") as f:
            f.write(data)

        if route == "pdf-to-word":
            return _to_docx(src)
        if route == "pdf-to-excel":
            return _to_xlsx(src)


# ---------------------------------------------------------------------------
# /upload/convert  — Office → PDF  (browser direct-upload, HMAC token auth)
# ---------------------------------------------------------------------------
@app.post("/upload/convert")
async def upload_convert(
    request: Request,
    route: str = Form(...),
    file: UploadFile = File(...),
):
    token = request.headers.get("x-dockdocs-upload-token", "")
    if not _verify_upload_token(token, route):
        raise HTTPException(status_code=401, detail="Invalid or expired upload token.")

    if route not in UPLOAD_ROUTES:
        raise HTTPException(
            status_code=422,
            detail=f"Unsupported route: {route}. Supported: {sorted(UPLOAD_ROUTES)}",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=422, detail="Empty file.")
    if len(data) > UPLOAD_MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large (max {UPLOAD_MAX_BYTES // 1024 // 1024} MB).",
        )

    filename = file.filename or _ROUTE_FILENAME[route]
    endpoint = _gotenberg_endpoint(route)

    # Build multipart for Gotenberg (internal network — no auth header needed)
    if route == "html-to-pdf":
        gotenberg_files = {"files": ("index.html", data, "text/html")}
        gotenberg_data = {}
    elif route == "pdf-to-pdfa":
        gotenberg_files = {"files": (filename, data, "application/pdf")}
        gotenberg_data = {"pdfa": "PDF/A-2b"}
    else:
        gotenberg_files = {"files": (filename, data, file.content_type or "application/octet-stream")}
        gotenberg_data = {}

    timeout = httpx.Timeout(connect=10.0, read=130.0, write=60.0, pool=5.0)
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(endpoint, files=gotenberg_files, data=gotenberg_data)
    except httpx.TimeoutException:
        logging.error("[upload/convert] Gotenberg timed out for route=%s size=%d", route, len(data))
        raise HTTPException(status_code=504, detail="Converter timed out — try a smaller file.")
    except httpx.RequestError as exc:
        logging.error("[upload/convert] Gotenberg unreachable: %s", exc)
        raise HTTPException(status_code=502, detail="Converter could not be reached.")

    if not resp.is_success:
        logging.error("[upload/convert] Gotenberg %d for route=%s", resp.status_code, route)
        raise HTTPException(status_code=502, detail=f"Converter returned {resp.status_code}.")

    pdf_bytes = resp.content
    if not pdf_bytes:
        raise HTTPException(status_code=502, detail="Converter returned an empty result.")

    logging.info("[upload/convert] OK route=%s size_in=%d size_out=%d", route, len(data), len(pdf_bytes))
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Cache-Control": "no-store"},
    )


# ---------------------------------------------------------------------------
# Helpers for /convert (PDF → office)
# ---------------------------------------------------------------------------
def _to_docx(src: str) -> Response:
    out = src.replace(".pdf", ".docx")
    cv = pdf2docx.Converter(src)
    cv.convert(out, start=0, end=None)
    cv.close()
    with open(out, "rb") as f:
        body = f.read()
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


def _to_xlsx(src: str) -> Response:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    with pdfplumber.open(src) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            if not tables:
                # No tables on this page — dump raw text as single column
                text = page.extract_text() or ""
                if text.strip():
                    ws = wb.create_sheet(title=f"Page {page_num}")
                    for line in text.split("\n"):
                        ws.append([line])
                continue
            for t_idx, table in enumerate(tables):
                sheet_name = f"P{page_num}" if len(tables) == 1 else f"P{page_num}-T{t_idx+1}"
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                for row in table:
                    ws.append([cell if cell is not None else "" for cell in row])

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/health")
def health():
    return {"ok": True}
